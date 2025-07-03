# -*- coding: utf-8 -*-
import os
import re
import json
import shutil
import requests
from openpyxl import load_workbook
from collections import defaultdict


def get_access_token():
    '''获取访问凭证'''
    url = 'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal'
    data = {
        "app_id": "xx",
        "app_secret": "xx"
    }
    ret = requests.post(url=url, data=json.dumps(data, ensure_ascii=False))
    data = ret.json()
    return data.get("tenant_access_token")

access_token = get_access_token()
headers = {
    "Authorization": f"Bearer {access_token}",
    "Content-Type": "application/json"
}


# 获取运费计算器全部sheet页数据，包含sheet_name与对应sheet_range
def get_sheet_info(spreadsheet_id):
    # https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/:spreadsheet_token/sheets/query
    url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{spreadsheet_id}/sheets/query"
    params = {
        "valueRenderOption": "ToString",
        "dateTimeRenderOption": "FormattedString"
    }
    res = requests.get(url, headers=headers, params=params)
    data = res.json()
    sheets_info = []
    for sheet in data['data']['sheets']:
        sheet_name = sheet['title']
        sheet_range = sheet['sheet_id']  # 从sheet_id中提取
        sheets_info.append({
            "sheet_name": sheet_name,
            "sheet_range": sheet_range
        })
    return sheets_info


def select_data(field_name, fieldvalue, app_token, table_id, view_id=None):
    '''多维表查询记录，根据多维表格中品名列与运费计算器中的产品名匹配'''
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/search"
    data = json.dumps({
        "filter": {
            "conjunction": "and",
            "conditions": [
                {
                    "field_name": field_name,
                    "operator": "is",
                    "value": [fieldvalue]
                }
            ]
        }
    }, ensure_ascii=False).encode('unicode_escape')

    response = requests.post(url, headers=headers, data=data)
    res_data = response.json()

    if res_data.get('code') == 0:
        items_list = res_data.get('data', {}).get('items', [])
        results = []
        for item in items_list:
            fields = item.get('fields', {})
            results.append({
                "Img_file_token": fields.get('图片', [{}])[0].get('file_token', ''),
                "Img_name": fields.get('M-SKU', [{}])[0].get('text', ''),
                "Chinese_name": fields.get('品名简称//——6/5', [{}])[0].get('text', ''),
                "English_name": fields.get('英文品名', [{}])[0].get('text', ''),
                "price": fields.get('进价：每件/套＄', ''),
                "Material": fields.get('材质', [{}])[0].get('text', ''),
                "HS_code": fields.get('HS编码', [{}])[0].get('text', ''),
                "Application": fields.get('用途', [{}])[0].get('text', ''),
                "brand": fields.get('品牌', [{}])[0].get('text', ''),
                "SKU": fields.get('SKU', [{}])[0].get('text', ''),
                "M_SKU": fields.get('M-SKU', [{}])[0].get('text', '')
            })
        return results


def get_reference_number(shipping_calculator_spreadsheet_id, sheet_name, base_code='G1235'):
    """
    生成reference_number
    :param sheet_name: sheet页名称，如 '5.16盈和加班美森IND9'
    :param base_code: 固定前缀，默认为'G1235'
    :return: 完整的reference_number，如'G123505020'前缀固定'G1235'+两位数月份+当月盈和序号
    """
    try:
        # 从sheet名称中提取月份
        month = sheet_name.split('.')[0]
        if not month.isdigit():
            raise ValueError(f"无法从sheet名称中提取月份: {sheet_name}")

        # 确保月份是两位数（个位数月份前面补0）
        month = month.zfill(2)  # 如果是"9"会变成"09"，如果是"11"保持不变

        # 创建一个列表来存储所有符合条件的sheet页
        sheet_list = []
        # 获取所有sheet信息
        sheets_info = get_sheet_info(shipping_calculator_spreadsheet_id)

        # 筛选符合条件的sheet页
        for info in sheets_info:
            # 修改筛选条件，使用实际的月份（可能是个位数或两位数）
            if f'{int(month)}.' in info['sheet_name'] and '盈和' in info['sheet_name'] and '一周' not in info['sheet_name']:
                sheet_list.append(info['sheet_name'])

        # 反转列表顺序
        sheet_list.reverse()

        # 创建序号映射
        sheet_dict = {name: idx + 1 for idx, name in enumerate(sheet_list)}

        # 获取当前sheet的序号
        current_number = sheet_dict.get(sheet_name)
        if current_number is None:
            raise ValueError(f"未找到对应的sheet页: {sheet_name}")

        # 生成三位数的序号
        sequence = str(current_number).zfill(3)

        # 组合最终的reference_number
        reference_number = f"{base_code}{month}{sequence}"

        return reference_number

    except Exception as e:
        print(f"生成reference_number时出错: {str(e)}")
        return None


def get_shipping_calculator_table(spreadsheet_id, range_):
    '''获取运费计算器云表格数据'''
    url = f'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_id}/values/{range_}'
    params = {
        "valueRenderOption": "ToString",
        "dateTimeRenderOption": "FormattedString"
    }
    response = requests.get(url, headers=headers, params=params)
    data = response.json()
    values = data.get('data', {}).get('valueRange', {}).get('values', [])
    product_name_list = []
    product_box_num_list = []
    box_size_list = []
    product_set_number_list = []
    real_weight_list = []
    # 过滤出指定列中值为指定值的行
    filtered_rows = [row for row in values if len(row) >= 7 and all(row[i] is not None for i in range(7)) and len(row) == 33]
    for product_info in filtered_rows[1::]:
        product_name = product_info[0].strip()
        product_set_number = product_info[1]
        box_size = product_info[2]
        product_box_num = product_info[3]
        real_weight = product_info[9]
        product_name_list.append(product_name)
        real_weight_list.append(real_weight)
        product_box_num_list.append(product_box_num)
        box_size_list.append(box_size)
        product_set_number_list.append(product_set_number)
    return product_name_list, product_box_num_list, real_weight_list, box_size_list, product_set_number_list


def get_fba_shipment_details_table(spreadsheet_id, range_, amazon_warehouse_code, sheet_name, declaration_quantity, M_SKU):
    '''支持分批发货记录的智能匹配'''
    # 初始化静态缓存
    if not hasattr(get_fba_shipment_details_table, '_cache'):
        get_fba_shipment_details_table._cache = {
            'stock_map': defaultdict(list),  # 库存记录缓存（按SKU+仓库）
            'usage_map': defaultdict(dict)  # 使用量跟踪（按表格ID+SKU+仓库）
        }

    # 生成缓存键
    cache_key = (spreadsheet_id, M_SKU, amazon_warehouse_code)

    # 获取表格数据（带缓存优化）
    if not get_fba_shipment_details_table._cache['stock_map'][cache_key]:
        url = f'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_id}/values/{range_}'
        params = {"valueRenderOption": "ToString", "dateTimeRenderOption": "FormattedString"}
        response = requests.get(url, headers=headers, params=params)
        data = response.json()
        values = data.get('data', {}).get('valueRange', {}).get('values', [])

        # 缓存有效记录
        if '加班美森' in sheet_name or '统配' in sheet_name or '限时达' in sheet_name:
            check_col = -2
        elif '普船' in sheet_name:
            check_col = -3
        elif '正班美森' in sheet_name:
            check_col = -1
        else:
            print('出现新的运输情况，请与仓库确认属于三种的哪种情况', sheet_name)

        valid_rows = [
            {
                "ShipmentID": row[0],
                "ReferenceID": row[1],
                "Declared_quantity": int(row[8]),
                "Amazon_warehouse_code": row[4],
                "M_SKU": row[7],
                "check_col": check_col
            }
            for row in values
            if row[7] == M_SKU
               and row[4] == amazon_warehouse_code
               and row[check_col] is not None
        ]
        get_fba_shipment_details_table._cache['stock_map'][cache_key] = valid_rows

        # 初始化使用量跟踪
        get_fba_shipment_details_table._cache['usage_map'][cache_key] = {
            'total_used': 0,
            'remaining': sum(r['Declared_quantity'] for r in valid_rows)
        }

    # 获取缓存数据
    stock_records = get_fba_shipment_details_table._cache['stock_map'][cache_key]
    usage_info = get_fba_shipment_details_table._cache['usage_map'][cache_key]
    current_qty = int(declaration_quantity)

    # 匹配逻辑
    results = []

    # 情况1：直接匹配
    for record in stock_records:
        if record['Declared_quantity'] == current_qty:
            results.append(record)
            break

    # 情况2：合并匹配
    if not results and usage_info['remaining'] >= current_qty:
        # 创建虚拟记录
        virtual_record = {
            **stock_records[0],  # 使用第一条有效记录
            "Declared_quantity": current_qty  # 显示实际分批发货量
        }
        results.append(virtual_record)

        # 更新使用量
        usage_info['total_used'] += current_qty
        usage_info['remaining'] -= current_qty

    return results


def get_fba_shipment_table(spreadsheet_id, range_, ShipmentID):
    '''下载FBA货件云表格数据'''
    url = f'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_id}/values/{range_}'
    params = {
        "valueRenderOption": "ToString",
        "dateTimeRenderOption": "FormattedString"
    }
    response = requests.get(url, headers=headers, params=params)
    data = response.json()
    values = data.get('data', {}).get('valueRange', {}).get('values', [])
    results = []
    # 过滤出指定列中值为指定值的行
    filtered_rows = [row for row in values if row[0] == ShipmentID]
    for product_info in filtered_rows:
        Delivery_address = product_info[5]
        country = product_info[6]
        results.append({
            "Delivery_address": Delivery_address,
            "country": country
        })
    return results


def get_current_box_num_List(product_name_list, product_box_num_list):
    current_box_num_List = []
    current_index = 1  # 当前填充的序号
    for product_name, box_num in zip(product_name_list, product_box_num_list):
        if ',' in product_name or '，' in product_name:  # 检查是否是混箱
            mixed_products = re.split(r'[，,]', product_name)  # 拆分混箱产品
            # 如果只有1个混箱，直接用前一个的序号
            if box_num == 1:
                current_box_num_List.append(str(current_index))
            else:
                range_str = f"{current_index}-{current_index + box_num - 1}"
                current_box_num_List.append(range_str)
                # 更新当前索引
            current_index += box_num
            # 为每个混箱产品添加重复的值
            current_box_num_List.extend([current_box_num_List[-1]] * (len(mixed_products) - 1))  # 重复最后一个结果，数量为混箱产品数量减去1
        else:
            # 单一产品，直接填充当前序号
            if box_num > 1:
                range_str = f"{current_index}-{current_index + box_num - 1}"
                current_box_num_List.append(range_str)
            else:
                current_box_num_List.append(str(current_index))  # 直接添加当前序号
            # 更新当前索引
            current_index += box_num
    return current_box_num_List


def write_yh_invoice(product_name_list, excel_file_path, product_info_list, amazon_warehouse_code, product_box_num_list, current_box_num_List, reference_number, delivery_address, country_code, city, continent, postcode):
    '''将产品信息和图片写入盈和发票Excel文件中'''
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active
    # 写入相同的数据项
    sheet['E3'] = 'YHE20210413024YHYB'
    sheet['E4'] = reference_number
    sheet['E6'] = amazon_warehouse_code
    sheet['E7'] = 'FBA地址'
    sheet['E8'] = 'Amazon'
    sheet['E9'] = 'Amazon'
    sheet['E10'] = delivery_address
    sheet['E12'] = city
    sheet['E13'] = continent
    sheet['E14'] = postcode
    sheet['E15'] = country_code
    sheet['E16'] = '13800138000'
    sheet['E17'] = '否'
    sheet['E19'] = '否'

    start_row = 23  # 起始行

    for current_row, info in enumerate(product_info_list, start=start_row):
        # 填写其他信息
        sheet[f'B{current_row}'] = info["product_box_num"]
        sheet[f'R{current_row}'] = info["HS_code"]
        sheet[f'C{current_row}'] = info["Chinese_name"]
        sheet[f'D{current_row}'] = info["English_name"]
        sheet[f'F{current_row}'] = info["calculate_declared_quantity"]
        sheet[f'E{current_row}'] = info["price"]
        sheet[f'G{current_row}'] = info["Material"]
        sheet[f'H{current_row}'] = info["Material"]
        sheet[f'S{current_row}'] = info["Application"]

    # 处理混箱合并单元格
    previous_box_number = None
    merge_start_row = None
    # 遍历所有行
    for current_row in range(start_row, start_row + len(product_info_list)):
        current_box_number = current_box_num_List[current_row - start_row]
        # 如果当前箱号与前一个箱号相同，则继续合并
        if current_box_number == previous_box_number:
            continue  # 继续检查下一行
        else:
            # 如果箱号变化，处理前一个合并区域
            if merge_start_row is not None and merge_start_row < current_row:
                # 合并所有相关列
                sheet.merge_cells(start_row=merge_start_row, start_column=2, end_row=current_row - 1, end_column=2)  # B列
            # 更新合并起始行和当前箱号
            merge_start_row = current_row
            previous_box_number = current_box_number
    # 循环结束后，处理最后一组连续相同箱号的行
    if merge_start_row is not None and merge_start_row < start_row + len(product_info_list):
        # 合并所有相关列
        sheet.merge_cells(start_row=merge_start_row, start_column=2, end_row=start_row + len(product_info_list) - 1, end_column=2)  # B列

        # 发票上半部分右上角按id分箱数S计算逻辑
        # 1. 生成子产品到原始product的索引映射
        expanded_indices = []
        for p_idx, product_name in enumerate(product_name_list):
            sub_products = product_name.split('，')  # 拆分混箱子产品
            expanded_indices.extend([p_idx] * len(sub_products))

        # 验证映射长度
        if len(expanded_indices) != len(product_info_list):
            raise ValueError(f"数据不匹配: product_info应有{len(expanded_indices)}条，实际{len(product_info_list)}条")

        # 2. 按ShipmentID和ReferenceID分组，并记录已处理的原始product索引
        shipment_groups = {}
        for info_idx, info in enumerate(product_info_list):
            p_idx = expanded_indices[info_idx]  # 当前子产品对应的原始product索引
            key = (info["ShipmentID"], info["ReferenceID"])
            box_num = product_box_num_list[p_idx]  # 原始product的箱数

            if key not in shipment_groups:
                shipment_groups[key] = {
                    "total_boxes": 0,
                    "processed_p_indices": set()  # 记录该组已处理的原始product索引
                }

            # 核心逻辑：每个原始product只在该Shipment组中累加一次箱数
            if p_idx not in shipment_groups[key]["processed_p_indices"]:
                shipment_groups[key]["total_boxes"] += box_num
                shipment_groups[key]["processed_p_indices"].add(p_idx)

        # 3. 将结果写入Excel（L列、M列、N列）
        write_row = 4
        for key in shipment_groups:
            shipment_id, ref_id = key
            total_boxes = shipment_groups[key]["total_boxes"]
            sheet[f'L{write_row}'] = shipment_id
            sheet[f'M{write_row}'] = ref_id
            sheet[f'N{write_row}'] = total_boxes
            write_row += 1

        workbook.save(excel_file_path)
        print(f"发票已生成: {excel_file_path}")


shipping_calculator_spreadsheet_id = 'xx' # 运费计算器id
multidimensional_table_token = 'xx' # 仓储多维表格token
multidimensional_table_id = 'xx' # 仓储多维表格id
multidimensional_table_view_id = 'xx' # 仓储多维表格view_id
fba_shipment_details_table_id = 'xx' # FBA表格id
fba_shipment_details_table_range = 'xx!A:O' # FBA货件明细表range
template_path = r'D:\work\data\发票\三个发票模板\盈和发票模板.xlsx' # 盈和发票模版地址
save_path = r'D:\work\data\发票\盈和' # 盈和生成发票保存地址


def run():
    try:
        sheets_info = get_sheet_info(shipping_calculator_spreadsheet_id) # 获取运费计算器全部sheet页数据，包含sheet_name与对应sheet_range
        all_product_info_lists = []
        for info in sheets_info:
            try:
                current_sheet_name = sheets_info[1]['sheet_name']
                current_date = re.search(r'(\d+\.\d+)', current_sheet_name).group(1)
                if current_date in info['sheet_name'] and '盈和' in info['sheet_name'] and '沃尔玛' not in info['sheet_name']:
                    sheet_name = info.get('sheet_name')
                    match_amazon_warehouse_code = re.findall(rf'{current_date}.*?(\w{{3}}\d)', sheet_name)
                    amazon_warehouse_code = match_amazon_warehouse_code[0]
                    sheet_range = info.get('sheet_range')
                    print(f"===================当前处理sheet页数据: {sheet_name}, Range: {sheet_range}, amazon_warehouse_code: {amazon_warehouse_code}===================")
                    product_name_list, product_box_num_list, real_weight_list, box_size_list, product_set_number_list = get_shipping_calculator_table(shipping_calculator_spreadsheet_id, sheet_range)
                    product_info_list = []
                    total_box_num = sum(product_box_num_list)
                    current_box_num_List = get_current_box_num_List(product_name_list, product_box_num_list)
                    print('product_name_list:{}，product_box_num_list:{}, current_box_num_List:{}，product_nums:{}，total_box_num:{}'.format(product_name_list, product_box_num_list, current_box_num_List, len(current_box_num_List), total_box_num))

                    for idx, (product_name, product_box_num, real_weight, box_size, product_set_number) in enumerate(
                            zip(product_name_list, product_box_num_list, real_weight_list, box_size_list,
                                product_set_number_list), start=1):
                        try:
                            if '，' in product_name or ',' in product_name:
                                mixed_products = re.split(r'[，,]', product_name)
                                print(f'混箱列表：{mixed_products}')
                                mixed_box_total = product_box_num  # 保存混箱的总箱数
                                for mixed_product in mixed_products:
                                    try:
                                        product_name_clean = mixed_product.split('x')[0].strip()
                                        product_num = mixed_product.split('x')[1].strip()
                                        calculate_declared_quantity = int(product_num) * int(product_box_num)
                                        print(f'处理其中的:{product_name_clean}')

                                        info_list = select_data('品名', product_name_clean, multidimensional_table_token, multidimensional_table_id, multidimensional_table_view_id)
                                        if not info_list:
                                            print(f"！！！！！！！！！！！！！！！！！！！！未找到产品详细信息: {product_name_clean}")
                                            continue

                                        for product_info in info_list:
                                            try:
                                                Chinese_name = product_info["Chinese_name"]
                                                English_name = product_info["English_name"]
                                                price = product_info["price"]
                                                Material = product_info["Material"]
                                                HS_code = product_info["HS_code"]
                                                Application = product_info["Application"]
                                                brand = product_info["brand"]
                                                M_SKU = product_info["M_SKU"]
                                                SKU = product_info["SKU"]

                                                fba_shipment_details = get_fba_shipment_details_table(fba_shipment_table_id, fba_shipment_details_table_range, amazon_warehouse_code, sheet_name, calculate_declared_quantity, M_SKU)
                                                if not fba_shipment_details:
                                                    print(f"！！！！！！！！！！！！！！！！！！！！未找到FBA货件明细: {product_name_clean}")
                                                    continue

                                                for details in fba_shipment_details:
                                                    try:
                                                        ShipmentID = details["ShipmentID"]
                                                        ReferenceID = details["ReferenceID"]
                                                        Declared_quantity = details["Declared_quantity"]

                                                        product_info = {
                                                            "ShipmentID": ShipmentID,
                                                            "ReferenceID": ReferenceID,
                                                            "Amazon_warehouse_code": amazon_warehouse_code,
                                                            "product_box_num": product_box_num,
                                                            "Chinese_name": Chinese_name,
                                                            "English_name": English_name,
                                                            "price": price,
                                                            "Declared_quantity": Declared_quantity,
                                                            "Material": Material,
                                                            "HS_code": HS_code,
                                                            "brand": brand,
                                                            "box_size": box_size,
                                                            "real_weight": real_weight,
                                                            "calculate_declared_quantity": calculate_declared_quantity,
                                                            "SKU": SKU,
                                                            "Application": Application,
                                                            "is_mixed": True,  # 添加混箱标记
                                                            "mixed_box_total": mixed_box_total,  # 添加混箱总箱数
                                                            "product_num": product_num  # 添加产品数量
                                                        }
                                                        product_info_list.append(product_info)
                                                    except Exception as e:
                                                        print(f"处理混箱货件明细时出错: {str(e)}")
                                                        continue
                                            except Exception as e:
                                                print(f"处理混箱产品信息时出错: {str(e)}")
                                                continue
                                    except Exception as e:
                                        print(f"处理混箱子产品时出错: {str(e)}")
                                        continue
                            else:
                                try:
                                    calculate_declared_quantity = int(product_set_number) * int(product_box_num)
                                    product_name = product_name.strip()
                                    if '×' in product_name:
                                        product_name = product_name.replace('×', '')
                                    print(f'正常单品单箱的:{product_name}')

                                    info_list = select_data('品名', product_name, multidimensional_table_token, multidimensional_table_id, multidimensional_table_view_id)
                                    if not info_list:
                                        print(f"！！！！！！！！！！！！！！！！！！！！未找到产品详细信息: {product_name}")
                                        continue

                                    for product_info in info_list:
                                        try:
                                            Chinese_name = product_info["Chinese_name"]
                                            English_name = product_info["English_name"]
                                            price = product_info["price"]
                                            Material = product_info["Material"]
                                            HS_code = product_info["HS_code"]
                                            Application = product_info["Application"]
                                            brand = product_info["brand"]
                                            M_SKU = product_info["M_SKU"]
                                            SKU = product_info["SKU"]

                                            fba_shipment_details = get_fba_shipment_details_table(fba_shipment_table_id, fba_shipment_details_table_range, amazon_warehouse_code, sheet_name, calculate_declared_quantity, M_SKU)
                                            if not fba_shipment_details:
                                                print(f"！！！！！！！！！！！！！！！！！！！！未找到FBA货件明细: {product_name}")
                                                continue

                                            for details in fba_shipment_details:
                                                try:
                                                    ShipmentID = details["ShipmentID"]
                                                    ReferenceID = details["ReferenceID"]
                                                    Declared_quantity = details["Declared_quantity"]

                                                    product_info = {
                                                        "ShipmentID": ShipmentID,
                                                        "ReferenceID": ReferenceID,
                                                        "Amazon_warehouse_code": amazon_warehouse_code,
                                                        "product_box_num": product_box_num,
                                                        "Chinese_name": Chinese_name,
                                                        "English_name": English_name,
                                                        "price": price,
                                                        "Declared_quantity": Declared_quantity,
                                                        "Material": Material,
                                                        "HS_code": HS_code,
                                                        "brand": brand,
                                                        "box_size": box_size,
                                                        "real_weight": real_weight,
                                                        "calculate_declared_quantity": calculate_declared_quantity,
                                                        "SKU": SKU,
                                                        "Application": Application,
                                                        "is_mixed": False  # 添加非混箱标记
                                                    }
                                                    product_info_list.append(product_info)
                                                except Exception as e:
                                                    print(f"处理单品货件明细时出错: {str(e)}")
                                                    continue
                                        except Exception as e:
                                            print(f"处理单品产品信息时出错: {str(e)}")
                                            continue
                                except Exception as e:
                                    print(f"处理单品时出错: {str(e)}")
                                    continue
                        except Exception as e:
                            print(f"处理产品 {product_name} 时出错: {str(e)}")
                            continue

                    # 处理地址信息
                    if product_info_list:
                        try:
                            first_shipment_id = product_info_list[0]["ShipmentID"]
                            address_info = get_fba_shipment_table(fba_shipment_table_id, fba_shipment_table_range, first_shipment_id)
                            if address_info:
                                Delivery_address = address_info[0]["Delivery_address"]
                                country_code = address_info[0]["country"]
                                try:
                                    delivery_address = Delivery_address.split('\n')[-1]
                                    Delivery_address_splited = delivery_address.split(',')
                                    city = Delivery_address_splited[1].strip() if len(Delivery_address_splited) > 1 else ""
                                    if len(Delivery_address_splited) > 2:
                                        continent_and_postcode = Delivery_address_splited[2].strip().split(' ')
                                        continent = continent_and_postcode[0].strip() if continent_and_postcode else ""
                                        postcode = continent_and_postcode[1].strip() if len(continent_and_postcode) > 1 else ""
                                    else:
                                        continent = ""
                                        postcode = ""
                                except Exception as e:
                                    print(f"处理地址信息时出错: {str(e)}")
                                    city = ""
                                    continent = ""
                                    postcode = ""
                            else:
                                delivery_address = ""
                                country_code = ""
                                city = ""
                                continent = ""
                                postcode = ""
                        except Exception as e:
                            print(f"获取地址信息时出错: {str(e)}")
                            delivery_address = ""
                            country_code = ""
                            city = ""
                            continent = ""
                            postcode = ""
                    else:
                        delivery_address = ""
                        country_code = ""
                        city = ""
                        continent = ""
                        postcode = ""

                    all_product_info_lists.append((sheet_name, product_info_list, amazon_warehouse_code))

                    try:
                        if len(current_box_num_List) != len(product_info_list):
                            print(len(current_box_num_List), current_box_num_List, len(product_info_list), product_info_list)
                            print(f"！！！！！！！！！！！！！！！！！！！！写入数量不一致，请手动排查: {sheet_name}！！！！！！！！！！！！！！！！！！！！")
                        else:
                            # 写入文件
                            reference_number = get_reference_number(shipping_calculator_spreadsheet_id, sheet_name)
                            modified_sheet_name = sheet_name.replace("盈和", f" {reference_number}盈和")
                            invoice_file_path = os.path.join(save_path, f'{modified_sheet_name}.xlsx')
                            shutil.copy(template_path, invoice_file_path)
                            write_yh_invoice(product_name_list, invoice_file_path, product_info_list, amazon_warehouse_code, product_box_num_list, current_box_num_List, reference_number, delivery_address, country_code, city, continent, postcode)
                    except Exception as e:
                        print(f"写入文件时出错: {str(e)}")

            except Exception as e:
                print(f"处理sheet {info['sheet_name']} 时出错: {str(e)}")
                continue

    except Exception as e:
        print(f"程序运行出错: {str(e)}")


if __name__ == '__main__':
    run()