# -*- coding: utf-8 -*-
import os
import re
import json
import shutil
import requests
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from collections import defaultdict


def get_access_token():
    '''获取访问凭证'''
    url = 'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal'
    data = {
        "app_id": "xx",
        "app_secret": "x"
    }
    ret = requests.post(url=url, data=json.dumps(data, ensure_ascii=False))
    data = ret.json()
    return data.get("tenant_access_token")

access_token = get_access_token()
headers = {
    "Authorization": f"Bearer {access_token}",
    "Content-Type": "application/json"
}

def get_sheet_info(spreadsheet_id):
    # https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/:spreadsheet_token/sheets/query
    url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{spreadsheet_id}/sheets/query"
    params = {
        "valueRenderOption": "ToString",
        "dateTimeRenderOption": "FormattedString"
    }
    res = requests.get(url, headers=headers, params=params)
    data = res.json()
    sheets_info = []
    for sheet in data['data']['sheets']:
        sheet_name = sheet['title']
        sheet_range = sheet['sheet_id']  # 从sheet_id中提取
        sheets_info.append({
            "sheet_name": sheet_name,
            "sheet_range": sheet_range
        })
    return sheets_info


def select_data(field_name, fieldvalue, app_token, table_id, view_id=None):
    '''多维表查询记录'''
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/search"
    data = json.dumps({
        "filter": {
            "conjunction": "and",
            "conditions": [
                {
                    "field_name": field_name,
                    "operator": "is",
                    "value": [fieldvalue]
                }
            ]
        }
    }, ensure_ascii=False).encode('unicode_escape')

    response = requests.post(url, headers=headers, data=data)
    res_data = response.json()

    if res_data.get('code') == 0:
        items_list = res_data.get('data', {}).get('items', [])
        results = []
        for item in items_list:
            fields = item.get('fields', {})
            results.append({
                "Img_file_token": fields.get('图片', [{}])[0].get('file_token', ''),
                "Img_name": fields.get('M-SKU', [{}])[0].get('text', ''),
                "Chinese_name": fields.get('品名简称//——6/5', [{}])[0].get('text', ''),
                "English_name": fields.get('英文品名', [{}])[0].get('text', ''),
                "price_rmb": fields.get('进价：每件/套￥', ''),
                "price": fields.get('进价：每件/套＄', ''),
                "Material": fields.get('材质', [{}])[0].get('text', ''),
                "HS_code": fields.get('HS编码', [{}])[0].get('text', ''),
                "Application": fields.get('用途', [{}])[0].get('text', ''),
                "brand": fields.get('品牌', [{}])[0].get('text', ''),
                "SKU": fields.get('SKU', [{}])[0].get('text', ''),
                "M_SKU": fields.get('M-SKU', [{}])[0].get('text', '')
            })
        return results


def get_shipping_calculator_table(spreadsheet_id, range_):
    '''下载运费计算器云表格数据'''
    url = f'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_id}/values/{range_}'
    params = {
        "valueRenderOption": "ToString",
        "dateTimeRenderOption": "FormattedString"
    }
    response = requests.get(url, headers=headers, params=params)
    data = response.json()
    values = data.get('data', {}).get('valueRange', {}).get('values', [])
    sell_product_code = values[-3][1]
    product_name_list = []
    real_weight_list = []
    product_box_num_list = []
    box_size_list = []
    product_set_number_list = []
    # 过滤出指定列中值为指定值的行
    filtered_rows = [row for row in values if len(row) >= 7 and all(row[i] is not None for i in range(7)) and len(row)== 33]
    for product_info in filtered_rows:
        product_name = product_info[0].strip()
        product_set_number = product_info[1]
        real_weight = product_info[4]
        box_size = product_info[2]
        product_box_num = product_info[3]
        product_name_list.append(product_name)
        real_weight_list.append(real_weight)
        product_box_num_list.append(product_box_num)
        box_size_list.append(box_size)
        product_set_number_list.append(product_set_number)
    return product_name_list, product_box_num_list, sell_product_code, real_weight_list, box_size_list, product_set_number_list


def get_fba_shipment_details_table(spreadsheet_id, range_, amazon_warehouse_code, sheet_name, declaration_quantity, M_SKU):
    '''支持分批发货记录的智能匹配'''
    # 初始化静态缓存
    if not hasattr(get_fba_shipment_details_table, '_cache'):
        get_fba_shipment_details_table._cache = {
            'stock_map': defaultdict(list),  # 库存记录缓存（按SKU+仓库）
            'usage_map': defaultdict(dict)  # 使用量跟踪（按表格ID+SKU+仓库）
        }

    # 生成缓存键
    cache_key = (spreadsheet_id, M_SKU, amazon_warehouse_code)

    # 获取表格数据（带缓存优化）
    if not get_fba_shipment_details_table._cache['stock_map'][cache_key]:
        url = f'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_id}/values/{range_}'
        params = {"valueRenderOption": "ToString", "dateTimeRenderOption": "FormattedString"}
        response = requests.get(url, headers=headers, params=params)
        data = response.json()
        values = data.get('data', {}).get('valueRange', {}).get('values', [])

        # 缓存有效记录
        if '加班美森' in sheet_name or '定提' in sheet_name:
            check_col = -2
        elif '普船' in sheet_name or '纽约卡派' in sheet_name or '萨凡纳' in sheet_name:
            check_col = -3
        elif '正班美森' in sheet_name:
            check_col = -1
        else:
            print('出现新的运输情况，请与仓库确认属于三种的哪种情况', sheet_name)

        valid_rows = [
            {
                "ShipmentID": row[0],
                "ReferenceID": row[1],
                "Declared_quantity": int(row[8]),
                "Amazon_warehouse_code": row[4],
                "M_SKU": row[7],
                "check_col": check_col
            }
            for row in values
            if row[7] == M_SKU
               and row[4] == amazon_warehouse_code
               and row[check_col] is not None
        ]
        get_fba_shipment_details_table._cache['stock_map'][cache_key] = valid_rows

        # 初始化使用量跟踪
        get_fba_shipment_details_table._cache['usage_map'][cache_key] = {
            'total_used': 0,
            'remaining': sum(r['Declared_quantity'] for r in valid_rows)
        }

    # 获取缓存数据
    stock_records = get_fba_shipment_details_table._cache['stock_map'][cache_key]
    usage_info = get_fba_shipment_details_table._cache['usage_map'][cache_key]
    current_qty = int(declaration_quantity)

    # 匹配逻辑
    results = []

    # 情况1：直接匹配
    for record in stock_records:
        if record['Declared_quantity'] == current_qty:
            results.append(record)
            break

    # 情况2：合并匹配
    if not results and usage_info['remaining'] >= current_qty:
        # 创建虚拟记录
        virtual_record = {
            **stock_records[0],  # 使用第一条有效记录
            "Declared_quantity": current_qty  # 显示实际分批发货量
        }
        results.append(virtual_record)

        # 更新使用量
        usage_info['total_used'] += current_qty
        usage_info['remaining'] -= current_qty

    return results


def download_img(file_token, token, img_name, index, save_path):
    '''下载图片并保存到指定目录'''
    headers = {
        "Authorization": f"Bearer {token}"
    }
    url = "https://open.feishu.cn/open-apis/drive/v1/medias/batch_get_tmp_download_url"
    params = {
        "file_tokens": file_token
    }
    response = requests.get(url, headers=headers, params=params)
    data = json.loads(response.text).get('data', {})
    tmp_download_urls = data.get('tmp_download_urls', [])

    if not tmp_download_urls:
        return

    download_img_url = tmp_download_urls[0].get('tmp_download_url')
    res = requests.get(download_img_url)

    img_file_name = f"{index} {img_name}.jpg"  # 包含序号的文件名
    img_file_path = os.path.join(save_path, img_file_name)

    with open(img_file_path, 'wb') as f:
        f.write(res.content)


def get_current_box_num_List(product_name_list, product_box_num_list):
    current_box_num_List = []
    current_index = 1  # 当前填充的序号
    for product_name, box_num in zip(product_name_list, product_box_num_list):
        if ',' in product_name or '，' in product_name:  # 检查是否是混箱
            mixed_products = re.split(r'[，,]', product_name)  # 拆分混箱产品
            # 如果只有1个混箱，直接用前一个的序号
            if box_num == 1:
                current_box_num_List.append(str(current_index))
            else:
                range_str = f"{current_index}-{current_index + box_num - 1}"
                current_box_num_List.append(range_str)
                # 更新当前索引
            current_index += box_num
            # 为每个混箱产品添加重复的值
            current_box_num_List.extend([current_box_num_List[-1]] * (len(mixed_products) - 1))  # 重复最后一个结果，数量为混箱产品数量减去1
        else:
            # 单一产品，直接填充当前序号
            if box_num > 1:
                range_str = f"{current_index}-{current_index + box_num - 1}"
                current_box_num_List.append(range_str)
            else:
                current_box_num_List.append(str(current_index))  # 直接添加当前序号
            # 更新当前索引
            current_index += box_num
    return current_box_num_List


def get_reference_number(shipping_calculator_spreadsheet_id, sheet_name, base_code='LL1235'):
    """
    生成reference_number
    :param sheet_name: sheet页名称，如 '11.1德速加班美森ABE8已上传系统' 或 '9.1德速加班美森ABE8已上传系统'
    :param base_code: 固定前缀，默认为'LL1235'
    :return: 完整的reference_number，如'LL123511001' 或 'LL123509001'
    """
    try:
        # 从sheet名称中提取月份
        month = sheet_name.split('.')[0]
        if not month.isdigit():
            raise ValueError(f"无法从sheet名称中提取月份: {sheet_name}")

        # 确保月份是两位数（个位数月份前面补0）
        month = month.zfill(2)  # 如果是"9"会变成"09"，如果是"11"保持不变

        # 创建一个列表来存储所有符合条件的sheet页
        sheet_list = []

        # 获取所有sheet信息
        sheets_info = get_sheet_info(shipping_calculator_spreadsheet_id)

        # 筛选符合条件的sheet页
        for info in sheets_info:
            # 修改筛选条件，使用实际的月份（可能是单位数或两位数）
            if f'{int(month)}.' in info['sheet_name'] and '德速' in info['sheet_name'] and '一周' not in info['sheet_name']:
                sheet_list.append(info['sheet_name'])

        # 反转列表顺序
        sheet_list.reverse()

        # 创建序号映射
        sheet_dict = {name: idx + 1 for idx, name in enumerate(sheet_list)}

        # 获取当前sheet的序号
        current_number = sheet_dict.get(sheet_name)
        if current_number is None:
            raise ValueError(f"未找到对应的sheet页: {sheet_name}")

        # 生成三位数的序号
        sequence = str(current_number).zfill(3)

        # 组合最终的reference_number
        reference_number = f"{base_code}{month}{sequence}"

        return reference_number

    except Exception as e:
        print(f"生成reference_number时出错: {str(e)}")
        return None


def write_yh_invoice(folder_path, excel_file_path, product_info_list, amazon_warehouse_code, sheet_name, current_box_num_List, total_box_num):
    '''将产品信息和图片写入德速发票Excel文件中'''
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # 对应sheet页填写reference_number
    reference_number = get_reference_number(shipping_calculator_spreadsheet_id, sheet_name)
    # 发票上半部分信息填写
    sheet['B2'] = reference_number
    sheet['B3'] = amazon_warehouse_code
    sheet['B4'] = '门到门'
    sheet['B10'] = 'U0001'
    sheet['B11'] = '普货（无任何电池）'
    sheet['G2'] = '美国'
    sheet['G3'] = '否'
    sheet['G4'] = '是'
    sheet['G9'] = total_box_num
    sheet['G10'] = 'USD'

    # 发票下半部分信息填写
    start_row = 13  # 起始行

    # 数据填充
    for current_row, info in enumerate(product_info_list, start=start_row):
        # 处理箱子序号
        current_box_number = current_box_num_List[current_row - start_row]

        sheet[f'A{current_row}'] = info.get("ShipmentID", "")
        sheet[f'B{current_row}'] = info.get("ReferenceID", "")
        sheet[f'C{current_row}'] = current_box_number
        sheet[f'D{current_row}'] = info.get("SKU", "")
        sheet[f'T{current_row}'] = info.get("real_weight", "")
        sheet[f'E{current_row}'] = info.get("English_name", "")
        sheet[f'F{current_row}'] = info.get("Chinese_name", "")
        sheet[f'G{current_row}'] = info.get("HS_code", "")
        sheet[f'H{current_row}'] = info.get("brand", "")
        sheet[f'I{current_row}'] = info.get("Material", "")
        sheet[f'J{current_row}'] = info.get("Application", "")

        # 获取 price_rmb 的值
        price_rmb = info.get("price_rmb", "")  # 如果不存在，默认为空字符串
        if type(price_rmb) != str:
            price_rmb = price_rmb * 1.2
        sheet[f'Q{current_row}'] = price_rmb

        # 获取 price 的值
        price = info.get("price", "")  # 如果不存在，默认为空字符串
        # print(type(price), price)
        if type(price) != str:
            price = price * 1.2
        sheet[f'R{current_row}'] = price

        # 处理箱子尺寸
        if '*' in info["box_size"]:
            length, width, height = info["box_size"].split('*')
            sheet[f'U{current_row}'] = length
            sheet[f'V{current_row}'] = width
            sheet[f'W{current_row}'] = height
        else:
            box_sizes = {'1号箱': (53, 29, 37), '2号箱': (53, 23, 29), '3号箱': (43, 21, 27), '4号箱': (35, 19, 23)}
            if info["box_size"] in box_sizes:
                sheet[f'U{current_row}'], sheet[f'V{current_row}'], sheet[f'W{current_row}'] = box_sizes[info["box_size"]]
        sheet[f'N{current_row}'] = info.get("Declared_quantity", "")

    previous_box_number = None
    merge_start_row = None

    # 遍历所有行
    for current_row in range(start_row, start_row + len(product_info_list)):
        current_box_number = current_box_num_List[current_row - start_row]

        # 如果当前箱号与前一个箱号相同，则继续合并
        if current_box_number == previous_box_number:
            continue  # 继续检查下一行
        else:
            # 如果箱号变化，处理前一个合并区域
            if merge_start_row is not None and merge_start_row < current_row:
                # 合并所有相关列
                sheet.merge_cells(start_row=merge_start_row, start_column=1, end_row=current_row - 1, end_column=1)  # A列
                sheet.merge_cells(start_row=merge_start_row, start_column=2, end_row=current_row - 1, end_column=2)  # B列
                sheet.merge_cells(start_row=merge_start_row, start_column=3, end_row=current_row - 1, end_column=3)  # C列
                sheet.merge_cells(start_row=merge_start_row, start_column=20, end_row=current_row - 1, end_column=20)  # T列
                sheet.merge_cells(start_row=merge_start_row, start_column=21, end_row=current_row - 1, end_column=21)  # U列
                sheet.merge_cells(start_row=merge_start_row, start_column=22, end_row=current_row - 1, end_column=22)  # V列
                sheet.merge_cells(start_row=merge_start_row, start_column=23, end_row=current_row - 1, end_column=23)  # W列

            # 更新合并起始行和当前箱号
            merge_start_row = current_row
            previous_box_number = current_box_number

    # 循环结束后，处理最后一组连续相同箱号的行
    if merge_start_row is not None and merge_start_row < start_row + len(product_info_list):
        # 合并所有相关列
        sheet.merge_cells(start_row=merge_start_row, start_column=1, end_row=start_row + len(product_info_list) - 1, end_column=1)  # A列
        sheet.merge_cells(start_row=merge_start_row, start_column=2, end_row=start_row + len(product_info_list) - 1, end_column=2)  # B列
        sheet.merge_cells(start_row=merge_start_row, start_column=3, end_row=start_row + len(product_info_list) - 1, end_column=3)  # C列
        sheet.merge_cells(start_row=merge_start_row, start_column=20, end_row=start_row + len(product_info_list) - 1, end_column=20)  # T列
        sheet.merge_cells(start_row=merge_start_row, start_column=21, end_row=start_row + len(product_info_list) - 1, end_column=21)  # U列
        sheet.merge_cells(start_row=merge_start_row, start_column=22, end_row=start_row + len(product_info_list) - 1, end_column=22)  # V列
        sheet.merge_cells(start_row=merge_start_row, start_column=23, end_row=start_row + len(product_info_list) - 1, end_column=23)  # W列

    # 插入图片
    img_files = {f.split(' ')[0]: f for f in os.listdir(folder_path) if f.endswith('.jpg')}

    for idx, info in enumerate(product_info_list, start=start_row):
        product_index = str(idx - start_row + 1)  # 获取产品在列表中的序号
        if product_index in img_files:
            img_file = img_files[product_index]
            img_path = os.path.join(folder_path, img_file)
            if not os.path.isfile(img_path):
                print(f"图片文件未找到: {img_path}")
                continue  # 找不到文件，跳过当前产品
            # 将图片插入到当前产品行
            img = Image(img_path)
            img.anchor = f'S{idx}'
            # 读取单元格的尺寸
            col_width_pixel = sheet.column_dimensions['S'].width * 7
            row_height_pixel = sheet.row_dimensions[idx].height or 15
            img.width = col_width_pixel
            img.height = row_height_pixel
            sheet.add_image(img)
        else:
            print(f"没有找到产品 {info.get('Chinese_name', '')} 对应的图片，保持该行为空，{sheet_name}")

    # 保存工作簿
    workbook.save(excel_file_path)
    print(f"产品信息和图片已写入到 {excel_file_path}")


shipping_calculator_spreadsheet_id = 'xx' # 运费计算器id
multidimensional_table_token = 'xx' # 仓储多维表格token
multidimensional_table_id = 'xx' # 仓储多维表格id
multidimensional_table_view_id = 'xx' # 仓储多维表格view_id
fba_shipment_details_table_id = 'xx' # FBA表格id
fba_shipment_details_table_range = 'xx!A:O' # FBA货件明细表range
template_path = r'D:\work\data\发票\三个发票模板\德速发票模版.xlsx' # 德速发票模版地址
save_path = r'D:\work\data\发票\德速' # 德速生成发票保存地址
save_image_path = r'D:\work\data\发票\德速\产品图片' # 德速发票图片保存地址


def run():
    try:
        sheets_info = get_sheet_info(shipping_calculator_spreadsheet_id)
        all_product_info_lists = []
        for info in sheets_info:
            try:
                current_sheet_name = sheets_info[1]['sheet_name']
                current_date = re.search(r'(\d+\.\d+)', current_sheet_name).group(1)
                if current_date in info['sheet_name'] and '德速' in info['sheet_name']:
                    sheet_name = info.get('sheet_name')
                    match_amazon_warehouse_code = re.findall(rf'{current_date}.*?(\w{{3}}\d)', sheet_name)
                    amazon_warehouse_code = match_amazon_warehouse_code[0]
                    sheet_range = info.get('sheet_range')
                    print(f"===================当前处理sheet页数据: {sheet_name}, Range: {sheet_range}, amazon_warehouse_code: {amazon_warehouse_code}===================")
                    product_name_list, product_box_num_list, sell_product_code, real_weight_list, box_size_list, product_set_number_list = get_shipping_calculator_table(shipping_calculator_spreadsheet_id, sheet_range)
                    product_info_list = []
                    total_box_num = sum(product_box_num_list)
                    current_box_num_List = get_current_box_num_List(product_name_list, product_box_num_list)
                    print('product_name_list:{}, product_box_num_list:{}, current_box_num_List:{}'.format(product_name_list, product_box_num_list, current_box_num_List))

                    for idx, (product_name, product_box_num, real_weight, box_size, product_set_number) in enumerate(zip(product_name_list, product_box_num_list, real_weight_list, box_size_list, product_set_number_list)):
                        try:
                            if '，' in product_name or ',' in product_name:
                                mixed_products = re.split(r'[，,]', product_name)
                                for mixed_product in mixed_products:
                                    try:
                                        product_name_clean = mixed_product.split('x')[0].strip()
                                        product_num = mixed_product.split('x')[1].strip()
                                        calculate_declared_quantity = int(product_num)
                                        declaration_quantity = int(product_num) * int(product_box_num)
                                        print(f'混箱中的:{product_name_clean}')

                                        if '×' in product_name_clean:
                                            product_name_clean = product_name_clean.replace('×', '')

                                        info_list = select_data('品名', product_name_clean, multidimensional_table_token, multidimensional_table_id, multidimensional_table_view_id)
                                        if not info_list:
                                            print(f"！！！！！！！！！！未找到产品详细信息: {product_name_clean}")
                                            continue

                                        for product_info in info_list:
                                            Img_file_token = product_info["Img_file_token"]
                                            Img_name = '混箱' + product_info["Img_name"]
                                            Chinese_name = product_info["Chinese_name"]
                                            English_name = product_info["English_name"]
                                            price_rmb = product_info["price_rmb"]
                                            price = product_info["price"]
                                            Material = product_info["Material"]
                                            HS_code = product_info["HS_code"]
                                            Application = product_info["Application"]
                                            brand = product_info["brand"]
                                            M_SKU = product_info["M_SKU"]
                                            SKU = product_info["SKU"]

                                            fba_shipment_details = get_fba_shipment_details_table(fba_shipment_details_table_id, fba_shipment_details_table_range, amazon_warehouse_code, sheet_name, declaration_quantity, M_SKU)
                                            if not fba_shipment_details:
                                                print(f"！！！！！！！！！！！！！！未找到FBA货件明细: {product_name_clean}, {M_SKU}")
                                                continue
                                            for details in fba_shipment_details:
                                                try:
                                                    ShipmentID = details["ShipmentID"]
                                                    ReferenceID = details["ReferenceID"]
                                                    Declared_quantity = details["Declared_quantity"]
                                                    # 创建目录用于保存图片
                                                    image_path = os.path.join(save_image_path, sheet_name)
                                                    os.makedirs(image_path, exist_ok=True)
                                                    # 保存图片
                                                    download_img(Img_file_token, access_token, Img_name, len(product_info_list) + 1, image_path)
                                                    product_info = {
                                                        "ShipmentID": ShipmentID,
                                                        "ReferenceID": ReferenceID,
                                                        "Amazon_warehouse_code": amazon_warehouse_code,
                                                        "product_box_num": product_box_num,
                                                        "Chinese_name": Chinese_name,
                                                        "English_name": English_name,
                                                        "price_rmb": price_rmb,
                                                        "price": price,
                                                        "Declared_quantity": Declared_quantity,
                                                        "SKU": SKU,
                                                        "M_SKU": M_SKU,
                                                        "Material": Material,
                                                        "HS_code": HS_code,
                                                        "brand": brand,
                                                        "box_size": box_size,
                                                        "real_weight": real_weight,
                                                        "Application": Application,
                                                        "product_num": product_num,
                                                        "declaration_quantity": declaration_quantity,
                                                        "calculate_declared_quantity": calculate_declared_quantity,
                                                        "is_mixed": True,  # 标记为混箱
                                                    }
                                                    product_info_list.append(product_info)
                                                except Exception as e:
                                                    print(f"处理货件明细时出错: {str(e)}")
                                                    continue
                                    except Exception as e:
                                        print(f"处理混箱产品时出错: {str(e)}")
                                        continue
                            else:
                                try:
                                    declaration_quantity = int(product_set_number) * int(product_box_num)
                                    product_name = product_name.strip()
                                    if '×' in product_name:
                                        product_name = product_name.replace('×', '')
                                    print(f'正常单品单箱的:{product_name}')
                                    info_list = select_data('品名', product_name, multidimensional_table_token, multidimensional_table_id, multidimensional_table_view_id)
                                    if not info_list:
                                        print(f"！！！！！！！！！！！！！！未找到产品详细信息: {product_name}")
                                        continue
                                    for product_info in info_list:
                                        Img_file_token = product_info["Img_file_token"]
                                        Img_name = product_info["Img_name"]
                                        Chinese_name = product_info["Chinese_name"]
                                        English_name = product_info["English_name"]
                                        price_rmb = product_info["price_rmb"]
                                        price = product_info["price"]
                                        Material = product_info["Material"]
                                        HS_code = product_info["HS_code"]
                                        Application = product_info["Application"]
                                        brand = product_info["brand"]
                                        M_SKU = product_info["M_SKU"]
                                        SKU = product_info["SKU"]

                                        fba_shipment_details = get_fba_shipment_details_table(fba_shipment_details_table_id, fba_shipment_details_table_range, amazon_warehouse_code, sheet_name, declaration_quantity, M_SKU)
                                        if not fba_shipment_details:
                                            print(f"！！！！！！！！！！！！！！未找到FBA货件明细: {product_name}, {M_SKU}")
                                            continue
                                        for details in fba_shipment_details:
                                            try:
                                                ShipmentID = details["ShipmentID"]
                                                ReferenceID = details["ReferenceID"]
                                                Declared_quantity = details["Declared_quantity"]
                                                # 创建目录用于保存图片
                                                image_path = os.path.join(save_image_path, sheet_name)
                                                os.makedirs(image_path, exist_ok=True)
                                                # 保存图片
                                                download_img(Img_file_token, access_token, Img_name, len(product_info_list) + 1, image_path)
                                                product_info = {
                                                    "ShipmentID": ShipmentID,
                                                    "ReferenceID": ReferenceID,
                                                    "Amazon_warehouse_code": amazon_warehouse_code,
                                                    "product_box_num": product_box_num,
                                                    "Chinese_name": Chinese_name,
                                                    "English_name": English_name,
                                                    "price_rmb": price_rmb,
                                                    "price": price,
                                                    "Declared_quantity": Declared_quantity,
                                                    "SKU": SKU,
                                                    "M_SKU": M_SKU,
                                                    "Material": Material,
                                                    "HS_code": HS_code,
                                                    "brand": brand,
                                                    "box_size": box_size,
                                                    "real_weight": real_weight,
                                                    "declaration_quantity": declaration_quantity,
                                                    "Application": Application
                                                }
                                                product_info_list.append(product_info)
                                            except Exception as e:
                                                print(f"处理货件明细时出错: {str(e)}")
                                                continue
                                except Exception as e:
                                    print(f"处理单品时出错: {str(e)}")
                                    continue
                        except Exception as e:
                            print(f"处理产品 {product_name} 时出错: {str(e)}")
                            continue
                    all_product_info_lists.append((sheet_name, product_info_list, amazon_warehouse_code, current_box_num_List))
                    try:
                        # 在这里检查数量一致性并写入文件
                        if len(current_box_num_List) != len(product_info_list):
                            print(len(current_box_num_List), len(product_info_list), product_info_list)
                            print(f"！！！！！！！！！！！！！！！！！！！！写入数量不一致，请手动排查: {sheet_name}！！！！！！！！！！！！！！！！！！！！")
                        else:
                            reference_number = get_reference_number(shipping_calculator_spreadsheet_id, sheet_name)
                            modified_sheet_name = sheet_name.replace("德速", f" {reference_number}德速")
                            invoice_file_path = os.path.join(save_path, f'{modified_sheet_name}.xlsx')
                            shutil.copy(template_path, invoice_file_path)
                            write_yh_invoice(os.path.join(save_image_path, sheet_name), invoice_file_path, product_info_list, amazon_warehouse_code, sheet_name, current_box_num_List, total_box_num)
                    except Exception as e:
                        print(f"写入文件时出错: {str(e)}")

            except Exception as e:
                print(f"处理sheet {info['sheet_name']} 时出错: {str(e)}")
                continue

    except Exception as e:
        print(f"程序运行出错: {str(e)}")


if __name__ == '__main__':
    run()
